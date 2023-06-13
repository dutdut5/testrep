import openpyxl as op
import pprint as pp


filename = 'Бланк заказа.xlsx'
wb = op.load_workbook(filename, data_only=True)
sheet = wb.active
subcat_dict = {}
max_rows = sheet.max_row
for i in range(7, max_rows+1):
    sku = sheet.cell(row=i, column=2).value
    subcat = sheet.cell(row=i, column=12).value
    if not sku:
        continue
    if subcat not in subcat_dict:
        subcat_dict[subcat] = [sku]
    else:
        subcat_dict[subcat].append(sku)


sortedict = dict(sorted(subcat_dict.items()))
print(sortedict)
with open('subcat.ini', 'w') as myfile:
    for key, value in subcat_dict.items():
        string_values = ', '.join(value)
        string_to_write = key + ' = ' + string_values + '\n'
        myfile.write(string_to_write)
